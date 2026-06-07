import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, Reference
import streamlit as st


# =========================
# STREAMLIT TITLE
# =========================

st.title("Diamond Price Validation Tool")

# =========================
# FILE UPLOAD
# =========================

uploaded_file = st.file_uploader(
    "Upload Excel File",
    type=["xlsx"]
)

# =========================
# START PROCESS
# =========================

if uploaded_file is not None:

    # =========================
    # READ EXCEL FILE
    # =========================

    df = pd.read_excel(uploaded_file)

    st.success("File Uploaded Successfully")

    OUTPUT_FILE = "diamond_price_validation_output.xlsx"

    # =========================
    # CLEAN COLUMNS
    # =========================

    df.columns = df.columns.str.strip()

    # =========================
    # CLARITY ORDER
    # =========================

    CLARITY_ORDER = [
        "FL",
        "IF",
        "VVS1",
        "VVS2",
        "VS1",
        "VS2",
        "SI1",
        "SI2",
        "I1",
        "I2",
        "I3"
    ]

    # =========================
    # COLOR ORDER
    # =========================

    COLOR_ORDER = list("DEFGHIJKLMNOPQRSTUVWXYZ")

    # =========================
    # RANK MAPS
    # =========================

    clarity_rank = {
        v: i for i, v in enumerate(CLARITY_ORDER)
    }

    color_rank = {
        v: i for i, v in enumerate(COLOR_ORDER)
    }

    # =========================
    # ADD ERROR COLUMNS
    # =========================

    df["Error Type"] = ""
    df["Error Message"] = ""

    # =========================
    # ERROR COUNTS
    # =========================

    clarity_error_count = 0
    color_error_count = 0

    # =====================================================
    # COLOR CHECK
    # SAME:
    # Shape + Size Grp + Clarity
    # =====================================================

    color_groups = df.groupby([
        "Shape",
        "Size Grp",
        "Clarity",
        "Quality"
    ])

    for group_name, group_df in color_groups:

        # ==============================================
        # GET MAX UPDATED PRICE FOR EACH COLOR
        # ==============================================

        color_price = (
            group_df.groupby("Color")["UPDATED PRICE"]
            .max()
            .reset_index()
        )

        # ==============================================
        # ADD COLOR RANK
        # ==============================================

        color_price["ColorRank"] = (
            color_price["Color"].map(color_rank)
        )

        # ==============================================
        # SORT BY COLOR
        # ==============================================

        color_price = color_price.sort_values(
            "ColorRank"
        )

        # ==============================================
        # COMPARE COLORS
        # ==============================================

        for i in range(len(color_price) - 1):

            current_row = color_price.iloc[i]
            next_row = color_price.iloc[i + 1]

            current_color = current_row["Color"]
            next_color = next_row["Color"]

            current_price = float(
                current_row["UPDATED PRICE"]
            )

            next_price = float(
                next_row["UPDATED PRICE"]
            )

            # ==========================================
            # SAME PRICE ALLOWED
            # ERROR ONLY IF LOWER COLOR IS HIGHER
            # ==========================================

            if next_price > current_price:

                error_rows = group_df[
                    group_df["Color"] == next_color
                ].index

                for idx in error_rows:

                    df.loc[
                        idx,
                        "Error Type"
                    ] += "Color Error | "

                    df.loc[
                        idx,
                        "Error Message"
                    ] += (
                        f"{next_color} color "
                        f"price higher than "
                        f"{current_color}. "
                    )

                    color_error_count += 1

    # =====================================================
    # CLARITY CHECK
    # SAME:
    # Shape + Size Grp + Color
    # =====================================================

    clarity_groups = df.groupby([
        "Shape",
        "Size Grp",
        "Color",
        "Quality"
    ])

    for group_name, group_df in clarity_groups:

        # ==============================================
        # GET MAX UPDATED PRICE FOR EACH CLARITY
        # ==============================================

        clarity_price = (
            group_df.groupby("Clarity")["UPDATED PRICE"]
            .max()
            .reset_index()
        )

        # ==============================================
        # ADD CLARITY RANK
        # ==============================================

        clarity_price["ClarityRank"] = (
            clarity_price["Clarity"].map(
                clarity_rank
            )
        )

        # ==============================================
        # SORT BY CLARITY
        # ==============================================

        clarity_price = clarity_price.sort_values(
            "ClarityRank"
        )

        # ==============================================
        # COMPARE CLARITY
        # ==============================================

        for i in range(len(clarity_price) - 1):

            current_row = clarity_price.iloc[i]
            next_row = clarity_price.iloc[i + 1]

            current_clarity = current_row["Clarity"]
            next_clarity = next_row["Clarity"]

            current_price = float(
                current_row["UPDATED PRICE"]
            )

            next_price = float(
                next_row["UPDATED PRICE"]
            )

            # ==========================================
            # SAME PRICE ALLOWED
            # ERROR ONLY IF LOWER CLARITY IS HIGHER
            # ==========================================

            if next_price > current_price:

                error_rows = group_df[
                    group_df["Clarity"] == next_clarity
                ].index

                for idx in error_rows:

                    df.loc[
                        idx,
                        "Error Type"
                    ] += "Clarity Error | "

                    df.loc[
                        idx,
                        "Error Message"
                    ] += (
                        f"{next_clarity} price "
                        f"higher than "
                        f"{current_clarity}. "
                    )

                    clarity_error_count += 1

    # =========================
    # CREATE OUTPUT EXCEL
    # =========================

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        # =====================
        # MAIN SHEET
        # =====================

        df.to_excel(
            writer,
            sheet_name="Checked_Data",
            index=False
        )

        # =====================
        # ONLY ERRORS SHEET
        # =====================

        error_df = df[
            df["Error Type"] != ""
        ]

        error_df.to_excel(
            writer,
            sheet_name="Only_Errors",
            index=False
        )

        # =====================
        # DASHBOARD SHEET
        # =====================

        dashboard_data = pd.DataFrame({

            "Type": [
                "Clarity Error",
                "Color Error",
                "Total Errors",
                "Total Checked Rows",
                "Correct Rows"
            ],

            "Count": [
                clarity_error_count,
                color_error_count,
                clarity_error_count + color_error_count,
                len(df),
                len(df) - (
                    clarity_error_count +
                    color_error_count
                )
            ]
        })

        dashboard_data.to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

    # =========================
    # LOAD WORKBOOK
    # =========================

    wb = load_workbook(OUTPUT_FILE)

    # =========================
    # STYLE MAIN SHEET
    # =========================

    ws = wb["Checked_Data"]

    red_fill = PatternFill(
        start_color="FF0000",
        end_color="FF0000",
        fill_type="solid"
    )

    blue_fill = PatternFill(
         start_color="0000FF",
         end_color="0000FF",
         fill_type="solid"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    # =========================
    # FIND ERROR COLUMN
    # =========================

    headers = [
        cell.value for cell in ws[1]
    ]

    error_col = (
        headers.index("Error Type") + 1
    )

    # =========================
    # HIGHLIGHT ERROR ROWS
    # =========================

    for row in range(2, ws.max_row + 1):

        error_value = ws.cell(
            row=row,
            column=error_col
        ).value

        if (
            error_value and
            str(error_value).strip() != ""
        ):

            for col in range(1, ws.max_column + 1):

                cell = ws.cell(
                    row=row,
                    column=col
                )

                cell.fill = red_fill
                cell.font = white_font

    # =========================
# BLUE HIGHLIGHT
# DIFFERENCE < 0
# AND NO OF DAYS <= 80
# =========================

    headers = [cell.value for cell in ws[1]]

    if (
        "DIFFERENCE" in headers
        and
        "No of Days" in headers
    ):
        diff_col = headers.index("DIFFERENCE") + 1
        days_col = headers.index("No of Days") + 1

        for row in range(2, ws.max_row + 1):

            try:

                diff_value = float(
                    ws.cell(
                        row=row,
                        column=diff_col
                    ).value
                )
                days_value = float(
                    ws.cell(
                        row=row,
                        column=days_col
                    ).value
                )
                if (
                    diff_value < 0
                    and
                    days_value <= 80
                    and
                    not ws.cell(row=row, column=error_col).value
                ):
                     for col in range(
                         1,
                         ws.max_column + 1
                     ):
                         cell = ws.cell(
                             row=row,
                             column=col
                         )
                         cell.fill = blue_fill
                         cell.font = white_font
                         
            except:
                pass
                        
               
                   

    

     

    
    # =========================
    # DASHBOARD CHART
    # =========================

    dashboard_ws = wb["Dashboard"]

    chart = BarChart()

    data = Reference(
        dashboard_ws,
        min_col=2,
        min_row=1,
        max_row=3
    )

    categories = Reference(
        dashboard_ws,
        min_col=1,
        min_row=2,
        max_row=3
    )

    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(categories)

    chart.title = "Diamond Validation Errors"

    chart.y_axis.title = "Count"

    chart.x_axis.title = "Error Type"

    dashboard_ws.add_chart(chart, "E2")

    # =========================
    # AUTO WIDTH
    # =========================

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value))
                if cell.value else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = length + 5

    # =========================
    # SAVE FILE
    # =========================

    wb.save(OUTPUT_FILE)

    # =========================
    # KPI DASHBOARD
    # =========================


    error_diamonds = len(
        df[
        df["Error Type"].str.strip() != ""
        ]
     )

    minus_count = 0

    if (
        "DIFFERENCE" in df.columns
        and
        "No of Days" in df.columns
    ):
        minus_count = len(
            df[
              (pd.to_numeric(df["DIFFERENCE"], errors="coerce") < 0)
               &
               (pd.to_numeric(df["No of Days"], errors="coerce") <= 80)
               &
               (df["Error Type"].str.strip() == "")
              ]
        )

    col1, col2, col3, col4, col5 = st.columns(5)

    col1.metric(
        "Total Diamonds",
        len(df)
    )

    col2.metric(
        "Error Diamonds",
        error_diamonds
    )

    col3.metric(
        "Clarity Errors",
        clarity_error_count
    )

    col4.metric(
        "Color Errors",
         color_error_count
    )

    col5.metric(
        "Negative Diff <=80 Days",
        minus_count
    )

    # =========================
    # ERROR CHART
    # =========================

    chart_df = pd.DataFrame({
        "Type": [
            "Clarity Errors",
            "Color Errors"
        ],
        "Count": [
            clarity_error_count,
            color_error_count
        ]
    })

    st.subheader("Error Summary")

    st.bar_chart(
        chart_df.set_index("Type")
    )

    negative_df = pd.DataFrame()

    if (
        "DIFFERENCE" in df.columns
        and
        "No of Days" in df.columns
    ):
        negative_df = df[
            (pd.to_numeric(df["DIFFERENCE"], errors="coerce") < 0)
            &
            (pd.to_numeric(df["No of Days"], errors="coerce") <= 80)
        ]

    if not negative_df.empty:
        st.subheader(
            "Negative Difference <= 80 Days"
        )
        st.dataframe(negative_df)


    # =========================
    # FILTERS
    # =========================

    st.subheader("Filters")
    
    selected_quality = st.multiselect(
        "Quality",
        sorted(df["Quality"].dropna().unique())
    )

    selected_shape = st.multiselect(
        "Shape",
        sorted(df["Shape"].dropna().unique())
    )

    selected_color = st.multiselect(
        "Color",
        sorted(df["Color"].dropna().unique())
    )

    selected_clarity = st.multiselect(
        "Clarity",
        sorted(df["Clarity"].dropna().unique())
    )

    selected_size = st.multiselect(
        "Size Grp",
        sorted(df["Size Grp"].dropna().unique())
    )

    selected_status = []
    if "Status" in df.columns:
        selected_status = st.multiselect(
            "Status",
             sorted(df["Status"].dropna().unique())
       )

    selected_lab = []
    if "Lab" in df.columns:
        selected_lab = st.multiselect(
            "Lab",
             sorted(df["Lab"].dropna().unique())
        )   



    

    filtered_df = df.copy()

    if selected_quality:
        filtered_df = filtered_df[
            filtered_df["Quality"].isin(selected_quality)
        ]

    if selected_shape:
        filtered_df = filtered_df[
            filtered_df["Shape"].isin(selected_shape)
        ]

    if selected_color:
        filtered_df = filtered_df[
            filtered_df["Color"].isin(selected_color)
        ]   

    if selected_clarity:
        filtered_df = filtered_df[
            filtered_df["Clarity"].isin(selected_clarity)
        ]

    if selected_size:
       filtered_df = filtered_df[
           filtered_df["Size Grp"].isin(selected_size)
       ]

    if selected_status:
       filtered_df = filtered_df[
           filtered_df["Status"].isin(selected_status)
       ]

    if selected_lab:
       filtered_df = filtered_df[
           filtered_df["Lab"].isin(selected_lab)
       ]


    st.write("Filtered Rows:", len(filtered_df))
    st.dataframe(filtered_df)

    # =========================
    # SELECT COLUMNS
    # =========================

    selected_columns = st.multiselect(
        "Columns To Download",
        filtered_df.columns.tolist(),
        default=filtered_df.columns.tolist()
    )

    if not selected_columns:
        st.warning("Please select at least one column")
        st.stop()

    download_df = filtered_df[selected_columns]

    st.dataframe(download_df)


  

        
        

    

    

    # =========================
    # DOWNLOAD BUTTON
    # =========================

    import io

    buffer = io.BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        download_df.to_excel(
            writer,
            index=False
        )

    st.download_button(
        label="Download Filtered Excel",
        data=buffer.getvalue(),
        file_name="filtered_diamonds.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
        

    st.success("Validation Completed Successfully")
